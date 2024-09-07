import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd


def read_excel_with_coordinates(file, x_start_row, x_end_row, x_column, y_start_row, y_end_row, y_column):
    workbook = load_workbook(file)
    sheet = workbook.active

    x_data = []
    y_data = []

    for row in range(x_start_row, x_end_row + 1):
        x_value = sheet.cell(row=row, column=x_column).value
        x_data.append(x_value)

    for row in range(y_start_row, y_end_row + 1):
        y_value = sheet.cell(row=row, column=y_column).value
        y_data.append(y_value)

    return x_data, y_data


def create_plot(x_data, y_data, chart_type="line", x_bar_name="x", y_bar_name="y", color="blue"):
    data = pd.DataFrame({x_bar_name: x_data, y_bar_name: y_data})

    plt.figure()

    sns.set_style("whitegrid")

    if chart_type == "line":
        sns.lineplot(data=data, x=x_bar_name, y=y_bar_name, color=color)
    elif chart_type == "bar":
        sns.barplot(data=data, x=x_bar_name, y=y_bar_name)
    elif chart_type == "scatter":
        sns.scatterplot(data=data, x=x_bar_name, y=y_bar_name)

    plt.title(f"{chart_type.capitalize()}")

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)

    return buffer
